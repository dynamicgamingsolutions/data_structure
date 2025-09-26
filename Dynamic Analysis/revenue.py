import os
import sys
dir_path = os.path.dirname(os.path.realpath(__file__))
sys.path.insert(0, os.path.join(dir_path, 'security'))
sys.path.insert(0, os.path.join(dir_path, 'sql'))
import calendar
from datetime import datetime, timedelta

import cipher
import keepass
import conn_module

import pymssql
import openpyxl
from openpyxl import Workbook

from flask import Flask, request, render_template, Blueprint

revenue_bp = Blueprint('revenue', __name__, static_folder='static', template_folder='templates')

@revenue_bp.after_request
def add_header(response):
    response.cache_control.no_store = True
    response.cache_control.no_cache = True
    response.cache_control.must_revalidate = True
    response.cache_control.proxy_revalidate = True
    response.cache_control.max_age = 0
    return response

@revenue_bp.route('/home')
def home():
    return render_template('revenue.html')

image_path = os.path.join(dir_path, 'security', 'The Man - The Myth - The Ray.jpg')

ascii_pass = cipher.get_user_comment(image_path)

mssql = keepass.credentials(ascii_pass)

@revenue_bp.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    file_path = os.path.join(dir_path, 'tmp', file.filename)
    file.save(file_path)

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        return 'Error loading workbook: ' + str(e)

    if 'Upload' in wb.sheetnames:
        sheet = wb['Upload']
        print(sheet.title)  # print the worksheet name to the terminal

        try:
            conn =conn_module.gather_conn(mssql[0], mssql[1], mssql[2], mssql[3])
            cursor = conn.cursor()

            # Iterate over the rows
            for row in sheet.iter_rows(min_row=2, values_only=True):  # assuming the first row contains headers
                # Create the INSERT INTO SQL query
                query = """INSERT INTO DGS_SLOT.dbo.revenue (serial_number, coin_in, promo, actual_win,days_on_floor, date, casino_id)
                            VALUES (%s, %s, %s,%s, %s, %s, %s )"""

                # Format the date
                if isinstance(row[5], datetime):
                    formatted_date = row[5].strftime('%Y-%m-%d')
                else:
                    formatted_date = row[5]  # if the date is not a datetime object, use it as is

                # Replace the original date with the formatted date
                row = row[:5] + (formatted_date,) + row[6:]

                # Execute the query with the values from the row
                cursor.execute(query, row)

            # Commit the transaction
            conn.commit()

        except Exception as e:
            return 'Error processing rows: ' + str(e)

        finally:
            conn.close()

        try:
            # Delete the file
            os.remove(file_path)
        except Exception as e:
            return 'Error deleting file: ' + str(e)

    else:
        return 'The workbook does not have a worksheet named "Upload".'

    return 'File uploaded successfully'

if __name__ == '__main__':
    app.run(debug=True)